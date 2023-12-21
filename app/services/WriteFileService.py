import os
from openpyxl import load_workbook
from datetime import datetime

class WriteFileService:
  _rootFolder = ""
  _dt = datetime.now()
  _ts = datetime.timestamp(_dt)

  _fileName = "formato"

  def __init__(self, rootFolder):
    self._rootFolder = rootFolder
  
  def writeFile(self):
    fileOriginPath = os.path.join(self._rootFolder, 'resources', "{}.xlsx".format(self._fileName))
    fileOutputPath = os.path.join(self._rootFolder, 'outputs', "{}.xlsx".format(self._fileName))

    print("1")
    wb = load_workbook(filename=fileOriginPath)
    print("2")
    ws = wb['Formulario220']
    print("3")
    ws['R10'] = 'Martinez'
    ws['X10'] = 'Morales'
    ws['AD10'] = 'Juan'
    ws['AI10'] = 'Camilo'
    print("4")
    wb.save(fileOutputPath)